"""Main application window for LibraryOfYore."""
import webbrowser
import datetime
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QComboBox, QListWidget, QListWidgetItem,
    QScrollArea, QGridLayout, QFrame, QMessageBox, QFileDialog,
    QStatusBar, QMenuBar, QMenu, QCheckBox, QGroupBox, QSplitter,
    QSystemTrayIcon, QApplication
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QAction, QFont, QIcon

from database.connection import test_connection, ensure_indexes
from database.models import NovelRepository, Novel
import api_server
from ui.setup_wizard import SetupWizard
from ui.add_novel_dialog import AddNovelDialog
from ui.novel_card import NovelCard
from config import STATUSES, EXPORTS_DIR, load_config, save_config, get_asset_path
from utils.helpers import bytes_to_pixmap

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class ExportWorker(QThread):
    """Background export to Excel."""
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, repo: NovelRepository, filepath: str):
        super().__init__()
        self.repo = repo
        self.filepath = filepath

    def run(self):
        try:
            data = self.repo.export_to_list()
            if not data:
                self.error.emit("No novels to export.")
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Novel Library"

            # Header
            headers = list(data[0].keys())
            ws.append(headers)
            header_fill = PatternFill(start_color="4caf50", end_color="4caf50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row in data:
                ws.append(list(row.values()))

            # Auto-width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(self.filepath)
            self.finished.emit(self.filepath)
        except Exception as e:
            self.error.emit(str(e))


class NovelRefreshWorker(QThread):
    """Background worker that re-scrapes Novelfire novels on startup to get the
    latest chapter count and status without blocking the UI."""

    novel_updated = pyqtSignal(str, int, str, str)  # novel_id, total_chapters, status, synopsis
    finished = pyqtSignal(int)                       # number of novels successfully refreshed

    def __init__(self, novels):
        super().__init__()
        self.novels = novels  # list of Novel objects with a novelfire source_url

    def run(self):
        from scrapers import get_scraper_for_url
        updated = 0
        for novel in self.novels:
            if not novel.source_url:
                continue
            scraper = get_scraper_for_url(novel.source_url)
            if not scraper:
                continue
            try:
                result = scraper.scrape(novel.source_url)
                if result.success:
                    total    = result.total_chapters if result.total_chapters else (novel.total_chapters or 0)
                    status   = result.status   if result.status   else novel.status
                    synopsis = result.synopsis if result.synopsis else novel.synopsis
                    self.novel_updated.emit(novel._id, total, status, synopsis)
                    updated += 1
            except Exception:
                pass  # Non-fatal — skip silently and move to next novel
        self.finished.emit(updated)


class MainWindow(QMainWindow):
    """Primary application window."""

    # Emitted from the API-server thread via api_server.set_progress_callback;
    # Qt routes it safely to the main thread.
    chapter_updated = pyqtSignal(str, int, int)  # novel_id, chapter, total

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Library of Yore")
        self.setWindowIcon(QIcon(get_asset_path("logo.ico")))
        self.resize(1300, 850)
        self.repo = None
        self.novel_cards = {}  # novel_id -> NovelCard
        self.current_sort = "last_read"
        self.grid_view = True
        self._force_quit = False  # True only when user picks Quit from tray

        self._check_db_connection()
        self._build_ui()
        self._apply_theme()
        self._setup_tray()
        self._refresh_library()
        self._start_novelfire_refresh()   # auto-fetch latest chapters/status on open

    def _check_db_connection(self):
        """Verify MongoDB on startup; show wizard if needed."""
        ok, msg = test_connection()
        if not ok:
            wizard = SetupWizard(self)
            if wizard.exec() != SetupWizard.DialogCode.Accepted:
                QMessageBox.critical(
                    self, "Cannot Start",
                    "MongoDB connection is required. The application will close."
                )
                import sys
                sys.exit(1)

        self.repo = NovelRepository()
        ensure_indexes()
        # Start local API server for the browser extension
        api_server.start()
        # Wire the extension → UI bridge: signal is thread-safe across Qt threads
        self.chapter_updated.connect(self._on_extension_chapter_update)
        api_server.set_progress_callback(
            lambda novel_id, chapter, total: self.chapter_updated.emit(novel_id, chapter, total or 0)
        )

    def _build_ui(self):
        # Menu bar
        menubar = self.menuBar()
        file_menu = menubar.addMenu("File")
        file_menu.addAction("Export to Excel", self._export_excel)
        file_menu.addSeparator()
        file_menu.addAction("Exit", self.close)

        view_menu = menubar.addMenu("View")
        view_menu.addAction("Grid View", lambda: self._set_view_mode(True))
        view_menu.addAction("List View", lambda: self._set_view_mode(False))
        view_menu.addSeparator()
        view_menu.addAction("Refresh", self._refresh_library)

        help_menu = menubar.addMenu("Help")
        help_menu.addAction("About", self._show_about)

        # Central widget with splitter
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        main_layout.addWidget(splitter)

        # === LEFT SIDEBAR ===
        sidebar = QWidget()
        sidebar.setMaximumWidth(260)
        sidebar.setMinimumWidth(200)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(12, 12, 12, 12)
        sidebar_layout.setSpacing(12)

        # Search
        search_box = QGroupBox("Search")
        search_layout = QVBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Title, author, notes...")
        self.search_input.textChanged.connect(self._refresh_library)
        search_layout.addWidget(self.search_input)
        search_box.setLayout(search_layout)
        sidebar_layout.addWidget(search_box)

        # Filters
        filter_box = QGroupBox("Filters")
        filter_layout = QVBoxLayout()
        self.status_checks = {}
        for status in STATUSES:
            cb = QCheckBox(status.title())
            cb.setChecked(True)
            cb.stateChanged.connect(self._refresh_library)
            self.status_checks[status] = cb
            filter_layout.addWidget(cb)
        filter_box.setLayout(filter_layout)
        sidebar_layout.addWidget(filter_box)

        # Sort
        sort_box = QGroupBox("Sort By")
        sort_layout = QVBoxLayout()
        self.sort_combo = QComboBox()
        self.sort_combo.addItems([
            "Last Read", "Title", "Rating", "Date Added", "Progress %"
        ])
        self.sort_combo.currentTextChanged.connect(self._on_sort_changed)
        sort_layout.addWidget(self.sort_combo)
        sort_box.setLayout(sort_layout)
        sidebar_layout.addWidget(sort_box)

        sidebar_layout.addStretch()

        # Stats
        self.stats_label = QLabel("0 novels")
        self.stats_label.setStyleSheet("color: #888; font-size: 11px;")
        sidebar_layout.addWidget(self.stats_label)

        splitter.addWidget(sidebar)

        # === RIGHT CONTENT ===
        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(16, 16, 16, 16)
        content_layout.setSpacing(12)

        # Top toolbar
        toolbar = QHBoxLayout()

        # Logo
        logo_label = QLabel()
        logo_path = get_asset_path("logo.ico")
        logo_pixmap = QIcon(logo_path).pixmap(40, 40)
        logo_label.setPixmap(logo_pixmap)
        toolbar.addWidget(logo_label)

        app_name = QLabel("Library of Yore")
        app_name.setStyleSheet("font-size: 18px; font-weight: bold; color: #d4af37;")
        toolbar.addWidget(app_name)

        toolbar.addSpacing(20)

        self.add_btn = QPushButton(" Add Novel")
        self.add_btn.setStyleSheet("padding: 8px 16px; font-weight: bold;")
        self.add_btn.clicked.connect(self._add_novel)
        toolbar.addWidget(self.add_btn)

        toolbar.addStretch()

        self.view_toggle_btn = QPushButton("List")
        self.view_toggle_btn.setCheckable(True)
        self.view_toggle_btn.clicked.connect(self._toggle_view)
        toolbar.addWidget(self.view_toggle_btn)

        self.export_btn = QPushButton("Export Excel")
        self.export_btn.clicked.connect(self._export_excel)
        toolbar.addWidget(self.export_btn)

        content_layout.addLayout(toolbar)

        # Scroll area for novel grid
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")

        self.grid_container = QWidget()
        self.grid_layout = QGridLayout(self.grid_container)
        self.grid_layout.setSpacing(16)
        self.grid_layout.setContentsMargins(8, 8, 8, 8)
        self.grid_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)

        self.scroll.setWidget(self.grid_container)
        content_layout.addWidget(self.scroll)

        splitter.addWidget(content)
        splitter.setSizes([250, 1050])

        # Status bar
        self.statusbar = QStatusBar()
        self.setStatusBar(self.statusbar)
        self.statusbar.showMessage(f"Ready  •  Browser extension API: localhost:{api_server.PORT}")

    def _apply_theme(self):
        """Apply dark theme stylesheet."""
        self.setStyleSheet("""
            QMainWindow { background-color: #121212; }
            QWidget { background-color: #121212; color: #e0e0e0; }
            QGroupBox {
                color: #d4af37;
                font-weight: bold;
                border: 1px solid #333;
                border-radius: 6px;
                margin-top: 10px;
                padding-top: 8px;
            }
            QGroupBox::title { subcontrol-origin: margin; left: 8px; padding: 0 4px; }
            QLineEdit {
                background-color: #1e1e1e;
                color: #eee;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 6px;
            }
            QPushButton {
                background-color: #2a2a2a;
                color: #eee;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #333; }
            QComboBox {
                background-color: #1e1e1e;
                color: #eee;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 4px;
            }
            QCheckBox { color: #ccc; }
            QCheckBox::indicator { width: 16px; height: 16px; }
            QScrollBar:vertical {
                background: #1a1a1a;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #444;
                border-radius: 6px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover { background: #555; }
            QMenuBar { background-color: #1a1a1a; color: #eee; }
            QMenuBar::item:selected { background-color: #333; }
            QMenu { background-color: #1a1a1a; color: #eee; border: 1px solid #444; }
            QMenu::item:selected { background-color: #333; }
            QStatusBar { background-color: #1a1a1a; color: #888; }
            QLabel { color: #ccc; }
        """)

    def _on_sort_changed(self, text: str):
        mapping = {
            "Last Read": "last_read",
            "Title": "title",
            "Rating": "rating",
            "Date Added": "date_added",
            "Progress %": "percent_complete",
        }
        self.current_sort = mapping.get(text, "last_read")
        self._refresh_library()

    def _set_view_mode(self, grid: bool):
        self.grid_view = grid
        self.view_toggle_btn.setText("⊞ Grid" if grid else "List")
        self._refresh_library()

    def _toggle_view(self):
        self.grid_view = not self.grid_view
        self._set_view_mode(self.grid_view)

    def _get_active_filters(self):
        statuses = [s for s, cb in self.status_checks.items() if cb.isChecked()]
        return statuses

    def _refresh_library(self):
        """Reload and display novels from database."""
        if not self.repo:
            return

        # Clear existing
        while self.grid_layout.count():
            item = self.grid_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self.novel_cards.clear()

        # Fetch
        search = self.search_input.text().strip()
        statuses = self._get_active_filters()
        novels = self.repo.get_all(
            status_filter=statuses if statuses else None,
            search_text=search,
            sort_by=self.current_sort,
            sort_order="desc"
        )

        self.stats_label.setText(f"{len(novels)} novel{'s' if len(novels) != 1 else ''}")
        self.statusbar.showMessage(f"Loaded {len(novels)} novels")

        if not novels:
            empty_label = QLabel("No novels found. Click 'Add Novel' to get started!")
            empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            empty_label.setStyleSheet("color: #666; font-size: 16px; margin-top: 50px;")
            self.grid_layout.addWidget(empty_label, 0, 0)
            return

        # Populate grid
        cols = 5 if self.grid_view else 1
        for i, novel in enumerate(novels):
            cover_bytes = None
            if novel.cover_image_id:
                cover_bytes = self.repo.get_cover(novel.cover_image_id)

            card = NovelCard(novel, cover_bytes)
            card.clicked.connect(self._on_card_clicked)
            card.edit_requested.connect(self._on_card_edit)
            card.delete_requested.connect(self._on_card_delete)
            card.open_url_requested.connect(self._open_url)
            card.chapter_plus.connect(self._on_chapter_plus)
            self.novel_cards[novel._id] = card

            row = i // cols
            col = i % cols
            self.grid_layout.addWidget(card, row, col)

    def _on_card_clicked(self, novel_id: str):
        self._on_card_edit(novel_id)

    def _on_card_edit(self, novel_id: str):
        novel = self.repo.get_by_id(novel_id)
        if not novel:
            return
        dialog = AddNovelDialog(self.repo, novel, self)
        dialog.novel_saved.connect(self._refresh_library)
        dialog.exec()

    def _on_card_delete(self, novel_id: str):
        reply = QMessageBox.question(
            self, "Confirm Delete",
            "Are you sure you want to delete this novel? This cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.repo.delete(novel_id)
            self._refresh_library()

    def _open_url(self, url: str):
        if url:
            webbrowser.open(url)

    def _on_extension_chapter_update(self, novel_id: str, chapter: int, total: int):
        """Slot called (on the main thread) when the browser extension updates progress."""
        if novel_id in self.novel_cards:
            self.novel_cards[novel_id].update_progress(chapter, total or None)
        self.statusbar.showMessage(
            f"Extension updated chapter → {chapter}", 3000
        )

    def _on_chapter_plus(self, novel_id: str):
        novel = self.repo.get_by_id(novel_id)
        if novel:
            novel.current_chapter += 1
            novel.last_read = datetime.datetime.utcnow()
            self.repo.update(novel)
            # Update card UI without full refresh
            if novel_id in self.novel_cards:
                self.novel_cards[novel_id].update_progress(
                    novel.current_chapter, novel.total_chapters
                )
            self.statusbar.showMessage(f"Updated '{novel.title}' to chapter {novel.current_chapter}", 3000)

    def _start_novelfire_refresh(self):
        """On startup, silently re-scrape all Novelfire novels to pull in the
        latest chapter count and status, then update the DB and each card."""
        if not self.repo:
            return
        all_novels = self.repo.get_all()
        novelfire_novels = [
            n for n in all_novels
            if n.source_url and "novelfire" in n.source_url.lower()
        ]
        if not novelfire_novels:
            return

        self.statusbar.showMessage(
            f"Auto-refreshing {len(novelfire_novels)} Novelfire novel(s) in background…"
        )
        self.refresh_worker = NovelRefreshWorker(novelfire_novels)
        self.refresh_worker.novel_updated.connect(self._on_novel_refreshed)
        self.refresh_worker.finished.connect(self._on_refresh_finished)
        self.refresh_worker.start()

    def _on_novel_refreshed(self, novel_id: str, total_chapters: int, status: str, synopsis: str):
        """Called (on main thread) for each novel the refresh worker finishes."""
        novel = self.repo.get_by_id(novel_id)
        if not novel:
            return

        changed = False
        if total_chapters and total_chapters != novel.total_chapters:
            novel.total_chapters = total_chapters
            changed = True
        if status and status != novel.status:
            novel.status = status
            changed = True
        if synopsis and synopsis != novel.synopsis:
            novel.synopsis = synopsis
            changed = True

        if changed:
            self.repo.update(novel)
            card = self.novel_cards.get(novel_id)
            if card:
                card.update_latest_chapter(novel.total_chapters or 0)
                card.update_status(novel.status)
                card.mark_updated()

    def _on_refresh_finished(self, count: int):
        msg = (
            f"Auto-refresh complete — {count} Novelfire novel(s) updated."
            if count else "Auto-refresh complete — no changes found."
        )
        self.statusbar.showMessage(msg, 6000)

    def _add_novel(self):
        dialog = AddNovelDialog(self.repo, parent=self)
        dialog.novel_saved.connect(self._refresh_library)
        dialog.exec()

    def _export_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Library",
            str(EXPORTS_DIR / "novel_library.xlsx"),
            "Excel Files (*.xlsx)"
        )
        if not path:
            return

        self.export_btn.setEnabled(False)
        self.export_btn.setText("⏳ Exporting...")

        self.export_worker = ExportWorker(self.repo, path)
        self.export_worker.finished.connect(self._on_export_done)
        self.export_worker.error.connect(self._on_export_error)
        self.export_worker.start()

    def _on_export_done(self, path: str):
        self.export_btn.setEnabled(True)
        self.export_btn.setText("Export Excel")
        msg = "Library exported to:" + chr(10) + path
        QMessageBox.information(self, "Export Complete", msg)

    def _on_export_error(self, msg: str):
        self.export_btn.setEnabled(True)
        self.export_btn.setText("Export Excel")
        QMessageBox.critical(self, "Export Failed", msg)

    def _show_about(self):
        QMessageBox.about(
            self, "About Library of Yore",
            "<h2>Library of Yore v1.0</h2>"
            "<p>A desktop bookmark tracker for web novels.</p>"
            "<p>Supports: Webnovel.com, Novelfire.net</p>"
            "<p>Built with Python, PyQt6, and MongoDB.</p>"
            f"<p><b>Browser Extension API:</b> localhost:{api_server.PORT}</p>"
        )

    def _setup_tray(self):
        """Create the system tray icon so the app keeps running when the window is closed."""
        if not QSystemTrayIcon.isSystemTrayAvailable():
            return

        self.tray_icon = QSystemTrayIcon(QIcon(get_asset_path("logo.ico")), self)
        self.tray_icon.setToolTip("Library of Yore — tracking in background")

        tray_menu = QMenu()

        open_action = QAction("Open Library", self)
        open_action.triggered.connect(self._show_window)
        tray_menu.addAction(open_action)

        tray_menu.addSeparator()

        quit_action = QAction("Quit", self)
        quit_action.triggered.connect(self._quit_app)
        tray_menu.addAction(quit_action)

        self.tray_icon.setContextMenu(tray_menu)
        # Single-click or double-click the tray icon → reopen window
        self.tray_icon.activated.connect(self._on_tray_activated)
        self.tray_icon.show()

    def _show_window(self):
        self.showNormal()
        self.raise_()
        self.activateWindow()

    def _on_tray_activated(self, reason):
        if reason in (QSystemTrayIcon.ActivationReason.Trigger,
                      QSystemTrayIcon.ActivationReason.DoubleClick):
            self._show_window()

    def _quit_app(self):
        self._force_quit = True
        QApplication.quit()

    def closeEvent(self, event):
        if self._force_quit:
            # Real quit — clean up and exit
            from database.connection import close_connection
            close_connection()
            event.accept()
        else:
            # Hide to tray instead of closing
            event.ignore()
            self.hide()
            if hasattr(self, "tray_icon"):
                self.tray_icon.showMessage(
                    "Library of Yore",
                    "Still tracking in the background. Right-click the tray icon to quit.",
                    QSystemTrayIcon.MessageIcon.Information,
                    3000
                )
